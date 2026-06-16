"""
JobSpy runner — scrapes job boards and saves results to Excel.

Usage:
    pip install -e .          # from repo root, first time only
    pip install openpyxl      # required for Excel export
    python run_scrape.py
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# Configuration (edit these values as needed)
# ---------------------------------------------------------------------------

SITE_NAMES = [
    "indeed",
    "linkedin",
    "google",
    "glassdoor",
    "naukri",
    "bayt",
    # zip_recruiter — US/Canada only, blocks India searches
    # bdjobs — Bangladesh-focused, not useful for Delhi NCR
]

SEARCH_TERM = "software engineer"
LOCATION = "Delhi NCR"
COUNTRY_INDEED = "india"

# Broader query — avoid remote/full-time/time filters that limit Google results
GOOGLE_SEARCH_TERM = "software engineer jobs in Delhi NCR India"

RESULTS_WANTED = 100

# Filters disabled for maximum results (set any of these to re-enable)
IS_REMOTE = False
JOB_TYPE = None       # e.g. "fulltime", "parttime", "contract", "internship"
HOURS_OLD = None      # e.g. 72 for last 3 days only

OUTPUT_FILE = Path("software_engineer_jobs_india.xlsx")
VERBOSE = 2  # 0 = errors only, 1 = errors + warnings, 2 = all logs

# Optional — helps when sites return 403 or CAPTCHA (Naukri, LinkedIn, Bayt)
PROXIES = None  # e.g. ["user:pass@host:port", "host:port"]
# LINKEDIN_FETCH_DESCRIPTION = True

# ---------------------------------------------------------------------------

MAX_COLUMN_WIDTH = 50
MIN_COLUMN_WIDTH = 10
DESCRIPTION_COLUMN_WIDTH = 50
NO_WRAP_COLUMNS = {"description"}


def timestamped_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def save_formatted_excel(jobs, output_file: Path) -> Path:
    """Write jobs to Excel with bold headers, wrapped cells, and readable column widths."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    temp_path = output_file.with_name(f"{output_file.stem}.tmp{output_file.suffix}")
    jobs.to_excel(temp_path, index=False, engine="openpyxl")

    workbook = load_workbook(temp_path)
    worksheet = workbook.active

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    header_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    header_alignment_no_wrap = Alignment(
        wrap_text=False, vertical="center", horizontal="center"
    )
    wrapped_alignment = Alignment(wrap_text=True, vertical="top")
    unwrapped_alignment = Alignment(wrap_text=False, vertical="top")

    column_names: dict[int, str] = {}
    for cell in worksheet[1]:
        column_names[cell.column] = str(cell.value or "").lower()
        is_no_wrap = column_names[cell.column] in NO_WRAP_COLUMNS
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment_no_wrap if is_no_wrap else header_alignment

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            col_name = column_names.get(cell.column, "")
            cell.alignment = (
                unwrapped_alignment
                if col_name in NO_WRAP_COLUMNS
                else wrapped_alignment
            )

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "").lower()

        if header in NO_WRAP_COLUMNS:
            worksheet.column_dimensions[column_letter].width = DESCRIPTION_COLUMN_WIDTH
            continue

        max_length = len(str(column_cells[0].value or ""))
        for cell in column_cells[1:]:
            if cell.value is not None:
                lines = str(cell.value).splitlines() or [""]
                max_length = max(max_length, max(len(line) for line in lines))

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, MIN_COLUMN_WIDTH),
            MAX_COLUMN_WIDTH,
        )

    worksheet.freeze_panes = worksheet["A2"]
    worksheet.sheet_view.topLeftCell = "A1"
    worksheet.row_dimensions[1].height = 24
    workbook.save(temp_path)

    try:
        shutil.move(str(temp_path), str(output_file))
        return output_file
    except PermissionError:
        fallback = timestamped_path(output_file)
        shutil.move(str(temp_path), str(fallback))
        print(
            f"\nCould not overwrite {output_file.name} — close it in Excel or your editor."
        )
        print(f"Saved to {fallback.resolve()} instead.")
        return fallback
    finally:
        temp_path.unlink(missing_ok=True)


def main() -> int:
    try:
        import pandas as pd
        from jobspy import scrape_jobs
    except ImportError as exc:
        print("Missing dependency. Install the project and openpyxl:")
        print("  pip install -e .")
        print("  pip install openpyxl")
        print(f"Error: {exc}")
        return 1

    print("JobSpy scrape starting...")
    print(f"  Sites:        {', '.join(SITE_NAMES)}")
    print(f"  Search:       {SEARCH_TERM!r}")
    print(f"  Location:     {LOCATION!r}")
    print(f"  Country:      {COUNTRY_INDEED}")
    print(f"  Results/site: {RESULTS_WANTED}")
    filters = []
    if IS_REMOTE:
        filters.append("remote")
    if JOB_TYPE:
        filters.append(f"job_type={JOB_TYPE}")
    if HOURS_OLD:
        filters.append(f"hours_old={HOURS_OLD}")
    filter_summary = ", ".join(filters) if filters else "none (all jobs)"
    print(f"  Filters:      {filter_summary}")
    print()

    scrape_kwargs = {
        "site_name": SITE_NAMES,
        "search_term": SEARCH_TERM,
        "location": LOCATION,
        "country_indeed": COUNTRY_INDEED,
        "results_wanted": RESULTS_WANTED,
        "verbose": VERBOSE,
    }

    if "google" in SITE_NAMES:
        scrape_kwargs["google_search_term"] = GOOGLE_SEARCH_TERM
    if IS_REMOTE:
        scrape_kwargs["is_remote"] = True
    if JOB_TYPE:
        scrape_kwargs["job_type"] = JOB_TYPE
    if HOURS_OLD is not None:
        scrape_kwargs["hours_old"] = HOURS_OLD
    if PROXIES:
        scrape_kwargs["proxies"] = PROXIES

    jobs = scrape_jobs(**scrape_kwargs)

    if jobs.empty:
        print("No jobs found. Try relaxing filters or adjusting the search term.")
        return 0

    print(f"\nFound {len(jobs)} jobs across {jobs['site'].nunique()} site(s).")
    print(jobs[["site", "title", "company", "location"]].head(10).to_string(index=False))
    print()

    try:
        saved_path = save_formatted_excel(jobs, OUTPUT_FILE)
    except ImportError:
        print("openpyxl is required for Excel export. Run: pip install openpyxl")
        return 1

    print(f"Saved to {saved_path.resolve()}")
    print(f"Finished at {datetime.now():%Y-%m-%d %H:%M:%S}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
